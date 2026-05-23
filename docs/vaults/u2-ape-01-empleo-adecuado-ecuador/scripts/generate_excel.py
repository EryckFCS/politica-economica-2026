import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference

def create_styled_excel():
    # 1. Definición de rutas
    script_dir = Path(__file__).resolve().parent
    vault_dir = script_dir.parent
    
    # Imagen de alta resolución a incrustar
    image_path = vault_dir / "assets" / "diagnostico_laboral.png"
    # Ruta de guardado final
    excel_path = vault_dir / "diagnostico_laboral_ecuador.xlsx"
    
    print(f"Ruta de la imagen de origen: {image_path}")
    print(f"Ruta del Excel de destino: {excel_path}")
    
    # 2. Inicializar libro de trabajo
    wb = openpyxl.Workbook()
    
    # Eliminar hoja default
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Hojas
    ws_datos = wb.create_sheet(title="Datos de Mercado Laboral")
    ws_graficos = wb.create_sheet(title="Gráficos Editables")
    ws_alta_calidad = wb.create_sheet(title="Visualización Alta Resolución")
    
    # Forzar mostrar líneas de cuadrícula en todas las hojas
    for ws in [ws_datos, ws_graficos, ws_alta_calidad]:
        ws.views.sheetView[0].showGridLines = True
        
    # --- ESTILOS GENERALES ---
    font_name = "Segoe UI"
    
    # Fuentes
    font_title = Font(name=font_name, size=16, bold=True, color="1F4E78")
    font_institution = Font(name=font_name, size=10, bold=True, color="595959")
    font_facultad = Font(name=font_name, size=9, italic=True, color="7F7F7F")
    font_tbl_title = Font(name=font_name, size=12, bold=True, color="1F4E78")
    font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    font_data = Font(name=font_name, size=11, color="000000")
    font_bold = Font(name=font_name, size=11, bold=True, color="000000")
    font_note = Font(name=font_name, size=9, italic=True, color="595959")
    
    # Fills (Zebra y cabeceras)
    fill_header1 = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Azul Profundo
    fill_header2 = PatternFill(start_color="A62626", end_color="A62626", fill_type="solid")  # Rojo Terracota
    fill_header3 = PatternFill(start_color="595959", end_color="595959", fill_type="solid")  # Gris Sólido
    
    fill_zebra1 = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")  # Zebra Azul
    fill_zebra2 = PatternFill(start_color="FDF2F2", end_color="FDF2F2", fill_type="solid")  # Zebra Roja
    fill_zebra3 = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")  # Zebra Gris
    
    # Bordes
    thin_side = Side(style='thin', color='D9D9D9')
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Alineaciones
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # --- HOJA 1: DATOS DE MERCADO LABORAL ---
    # Cabecera institucional
    ws_datos["B2"] = "UNIVERSIDAD NACIONAL DE LOJA"
    ws_datos["B2"].font = font_institution
    
    ws_datos["B3"] = "Facultad Jurídica, Social y Administrativa - Carrera de Economía"
    ws_datos["B3"].font = font_facultad
    
    ws_datos["B4"] = "Diagnóstico Estructural del Mercado Laboral Ecuatoriano (2024)"
    ws_datos["B4"].font = font_title
    
    ws_datos["B5"] = "Datos del Práctico Experimental de Política Económica - IV Trimestre 2024"
    ws_datos["B5"].font = font_note
    
    # --- TABLA 1: Estructura del Mercado Laboral (Col B y C) ---
    ws_datos["B7"] = "Estructura del Mercado Laboral Ecuatoriano (2024)"
    ws_datos["B7"].font = font_tbl_title
    
    ws_datos["B8"] = "Categoría"
    ws_datos["C8"] = "Porcentaje de la PEA (%)"
    for col in ["B8", "C8"]:
        ws_datos[col].font = font_header
        ws_datos[col].fill = fill_header1
        ws_datos[col].alignment = align_center
        ws_datos[col].border = border_thin
        
    data_t1 = [
        ("Empleo Adecuado", 0.359),
        ("Otro Inadecuado", 0.394),
        ("Subempleo", 0.210),
        ("Desempleo", 0.037)
    ]
    
    row_idx = 9
    for i, (cat, val) in enumerate(data_t1):
        ws_datos.cell(row=row_idx, column=2, value=cat).alignment = align_left
        ws_datos.cell(row=row_idx, column=3, value=val).alignment = align_right
        
        # Formatos
        ws_datos.cell(row=row_idx, column=2).font = font_data
        ws_datos.cell(row=row_idx, column=3).font = font_bold
        ws_datos.cell(row=row_idx, column=3).number_format = '0.0%'
        
        # Cebras y Bordes
        fill = fill_zebra1 if i % 2 == 1 else PatternFill(fill_type=None)
        for col_idx in [2, 3]:
            cell = ws_datos.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            if fill.fill_type:
                cell.fill = fill
        row_idx += 1
        
    # Fila de Total (Opcional, suma de los tres subempleos + desempleo = 100%)
    ws_datos.cell(row=row_idx, column=2, value="Población Económicamente Activa (PEA)").alignment = align_left
    ws_datos.cell(row=row_idx, column=2).font = font_bold
    ws_datos.cell(row=row_idx, column=2).border = border_thin
    
    total_cell = ws_datos.cell(row=row_idx, column=3, value="=SUM(C9:C12)")
    total_cell.alignment = align_right
    total_cell.font = font_bold
    total_cell.number_format = '0.0%'
    total_cell.border = border_thin
    
    # --- TABLA 2: Tasa de Informalidad Laboral (Col E y F) ---
    ws_datos["E7"] = "Tasa de Informalidad Laboral por Área Geográfica"
    ws_datos["E7"].font = font_tbl_title
    
    ws_datos["E8"] = "Área Geográfica"
    ws_datos["F8"] = "Tasa de Informalidad (%)"
    for col in ["E8", "F8"]:
        ws_datos[col].font = font_header
        ws_datos[col].fill = fill_header2
        ws_datos[col].alignment = align_center
        ws_datos[col].border = border_thin
        
    data_t2 = [
        ("Área Rural", 0.758),
        ("Nivel Nacional", 0.552),
        ("Área Urbana", 0.436)
    ]
    
    row_idx2 = 9
    for i, (area, val) in enumerate(data_t2):
        ws_datos.cell(row=row_idx2, column=5, value=area).alignment = align_left
        ws_datos.cell(row=row_idx2, column=6, value=val).alignment = align_right
        
        # Formatos
        ws_datos.cell(row=row_idx2, column=5).font = font_data
        ws_datos.cell(row=row_idx2, column=6).font = font_bold
        ws_datos.cell(row=row_idx2, column=6).number_format = '0.0%'
        
        # Cebras y Bordes
        fill = fill_zebra2 if i % 2 == 1 else PatternFill(fill_type=None)
        for col_idx in [5, 6]:
            cell = ws_datos.cell(row=row_idx2, column=col_idx)
            cell.border = border_thin
            if fill.fill_type:
                cell.fill = fill
        row_idx2 += 1
        
    # --- TABLA 3: Indicadores Adicionales y Brechas (Col B y C abajo) ---
    start_r3 = 16
    ws_datos.cell(row=start_r3, column=2, value="Indicadores de Brechas Laborales y Desempleo").font = font_tbl_title
    
    ws_datos.cell(row=start_r3+1, column=2, value="Métrica / Brecha Diagnosticada").font = font_header
    ws_datos.cell(row=start_r3+1, column=2).fill = fill_header3
    ws_datos.cell(row=start_r3+1, column=2).alignment = align_center
    ws_datos.cell(row=start_r3+1, column=2).border = border_thin
    
    ws_datos.cell(row=start_r3+1, column=3, value="Valor (%)").font = font_header
    ws_datos.cell(row=start_r3+1, column=3).fill = fill_header3
    ws_datos.cell(row=start_r3+1, column=3).alignment = align_center
    ws_datos.cell(row=start_r3+1, column=3).border = border_thin
    
    data_t3 = [
        ("Empleo Adecuado - Hombres", 0.414),
        ("Empleo Adecuado - Mujeres", 0.284),
        ("Empleo Adecuado - Área Rural", 0.194),
        ("Tasa de Desempleo - Nacional", 0.037),
        ("Tasa de Desempleo - Área Urbana", 0.050),
        ("Tasa de Desempleo - Área Rural", 0.014)
    ]
    
    row_idx3 = start_r3 + 2
    for i, (metric, val) in enumerate(data_t3):
        ws_datos.cell(row=row_idx3, column=2, value=metric).alignment = align_left
        ws_datos.cell(row=row_idx3, column=3, value=val).alignment = align_right
        
        # Formatos
        ws_datos.cell(row=row_idx3, column=2).font = font_data
        ws_datos.cell(row=row_idx3, column=3).font = font_bold
        ws_datos.cell(row=row_idx3, column=3).number_format = '0.0%'
        
        # Cebras y Bordes
        fill = fill_zebra3 if i % 2 == 1 else PatternFill(fill_type=None)
        for col_idx in [2, 3]:
            cell = ws_datos.cell(row=row_idx3, column=col_idx)
            cell.border = border_thin
            if fill.fill_type:
                cell.fill = fill
        row_idx3 += 1
        
    # Nota de Fuente al final
    ws_datos.cell(row=row_idx3+1, column=2, value="Nota: Datos oficiales de la ENEMDU publicados por el Instituto Nacional de Estadística y Censos (INEC) para el cierre de 2024.").font = font_note
    ws_datos.cell(row=row_idx3+2, column=2, value="Elaborado para distribución grupal académica. Todos los datos están enlazados dinámicamente con los gráficos en la siguiente hoja.").font = font_note
    
    # Autoajustar anchos de columna en Hoja 1
    for col in ws_datos.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Ignorar fila de título grande B4
        for cell in col:
            if cell.row in [4, row_idx3+1, row_idx3+2]:
                continue
            val_str = str(cell.value or '')
            if cell.number_format == '0.0%' and isinstance(cell.value, float):
                val_str = f"{cell.value * 100:.1f}%"
            max_len = max(max_len, len(val_str))
        ws_datos.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws_datos.column_dimensions["A"].width = 3 # Margen izquierdo
    ws_datos.column_dimensions["D"].width = 4 # Separador de tablas
    
    
    # --- HOJA 2: GRÁFICOS EDITABLES ---
    ws_graficos["B2"] = "GRÁFICOS OPERATIVOS DEL MERCADO LABORAL"
    ws_graficos["B2"].font = font_title
    ws_graficos["B3"] = "Gráficos de Excel nativos y 100% editables para presentaciones y trabajos grupales."
    ws_graficos["B3"].font = font_note
    
    # Gráfico 1: Estructura del Mercado Laboral (BarChart Vertical)
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Estructura del Mercado Laboral Ecuatoriano (2024)"
    chart1.y_axis.title = "Porcentaje de la PEA (%)"
    chart1.x_axis.title = "Categoría de Empleo"
    
    # Referencias a la Hoja 1
    # data1 incluye la cabecera "Porcentaje de la PEA (%)"
    data1 = Reference(ws_datos, min_col=3, min_row=8, max_row=12)
    cats1 = Reference(ws_datos, min_col=2, min_row=9, max_row=12)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.legend = None # Sin leyenda porque solo hay una serie y los nombres están en el eje X
    chart1.width = 15
    chart1.height = 10
    
    # Añadir al gráfico
    ws_graficos.add_chart(chart1, "B5")
    
    # Gráfico 2: Tasa de Informalidad (BarChart Horizontal)
    chart2 = BarChart()
    chart2.type = "bar" # Barra horizontal
    chart2.style = 13
    chart2.title = "Tasa de Informalidad Laboral por Área Geográfica (2024)"
    chart2.x_axis.title = "Área Geográfica"
    chart2.y_axis.title = "Tasa de Informalidad (%)"
    
    data2 = Reference(ws_datos, min_col=6, min_row=8, max_row=11)
    cats2 = Reference(ws_datos, min_col=5, min_row=9, max_row=11)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.legend = None
    chart2.width = 15
    chart2.height = 10
    
    ws_graficos.add_chart(chart2, "K5")
    
    
    # --- HOJA 3: VISUALIZACIÓN DE ALTA RESOLUCIÓN ---
    ws_alta_calidad["B2"] = "DIAGNÓSTICO ACADÉMICO - MATPLOTLIB VECTORIAL"
    ws_alta_calidad["B2"].font = font_title
    ws_alta_calidad["B3"] = "Esta hoja incrusta la visualización oficial generada para el documento académico (index.pdf)."
    ws_alta_calidad["B3"].font = font_note
    ws_alta_calidad["B4"] = "Ideal para copiar directamente como imagen de alta resolución para diapositivas de sustentación."
    ws_alta_calidad["B4"].font = font_note
    
    if image_path.exists():
        try:
            img = Image(str(image_path))
            # Ajustar tamaño de imagen para que no tape todo
            img.width = 750
            img.height = 317
            ws_alta_calidad.add_image(img, "B6")
            print("Imagen de alta calidad incrustada exitosamente en la hoja 3.")
        except Exception as e:
            print(f"Error incrustando la imagen en Excel: {e}")
            ws_alta_calidad["B6"] = f"[Error cargando la imagen: {e}]"
    else:
        print(f"Advertencia: No se encontró la imagen en {image_path}. Verifica que generate_plots.py haya corrido.")
        ws_alta_calidad["B6"] = "[Imagen de alta resolución no encontrada - Asegúrate de correr scripts/generate_plots.py primero]"
        
    # Guardar archivo
    wb.save(str(excel_path))
    print(f"Archivo Excel generado exitosamente en: {excel_path}")

if __name__ == "__main__":
    create_styled_excel()
